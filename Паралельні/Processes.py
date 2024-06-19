from multiprocessing import Process, Queue
import time
import random
from openpyxl import Workbook, load_workbook
def average(arr, process_index, result_queue):
    start_time = time.time()
    time.sleep(1)
    total = sum(arr)
    length = len(arr)
    average_val = total / length
    print("Process №", process_index)
    print(average_val)
    end_time = time.time()
    execution_time = end_time - start_time
    print("Execution time for part:", round(execution_time, 5), "seconds")
    result_queue.put(round(execution_time, 6))

def split_arr(arr, size_count):
    part_size = len(arr) // size_count
    result = [arr[i * part_size:(i + 1) * part_size] for i in range(size_count)]
    return result

def created_processes(splited_arr, result_queue):
    processes = [Process(target=average, args=(sub_arr, i + 1, result_queue)) for i, sub_arr in enumerate(splited_arr)]
    return processes


if __name__ == "__main__":
    count_part = int(input("Введіть кількість процесів: "))
    try:
        arr_length = int(input("Введіть розмірність масиву, яка ділиться на {}: ".format(count_part)))
    except ValueError:
        print("Введіть ціле число")
        exit()
    if arr_length % count_part != 0:
        print("Масив повинен бути кратним ", count_part)
        exit()
    arr = [random.randint(100, 100000) for _ in range(arr_length)]
    start_time = time.time()
    splited_arr = split_arr(arr, count_part)
    result_queue = Queue()
    processes = created_processes(splited_arr, result_queue)
    for proc in processes:
        proc.start()

    for proc in processes:
        proc.join()

    results = []
    while not result_queue.empty():
        results.append(result_queue.get())

    end_time = time.time()
    execution_time = end_time - start_time
    print("Execution time for all program:", execution_time, "seconds")
    try:
        Excell = load_workbook("result_multe.xlsx")
    except FileNotFoundError:
        Excell = Workbook()
    Ex_sheet = Excell.active
    col = 13

    i = 1
    Ex_sheet.cell(row=1, column=col).value = "Час витрачений на процес"
    while i <= count_part:
        Ex_sheet.cell(row=i + 1, column=col).value = results[i - 1]
        i += 1
    Ex_sheet.cell(row=count_part + 2, column=col).value = "Весь час загалом на програму"
    Ex_sheet.cell(row=count_part + 3, column=col).value = execution_time
    Excell.save("result_multe.xlsx")
